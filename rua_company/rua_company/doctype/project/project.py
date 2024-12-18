# Copyright (c) 2024, Yamen Zakhour and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import flt
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import base64
from io import BytesIO
import json
import re
import os
from PIL import Image as PILImage
import io
from . import project_calculations

class Project(Document):
    def validate(self):
        # Handle serial number for In Progress projects
        if self.status == "In Progress" and self.serial_number == 0:
            # Get the highest serial number
            highest_serial = frappe.db.get_value(
                "Project",
                {"serial_number": (">", 0)},
                "serial_number",
                order_by="serial_number desc"
            )
            
            # Assign next serial number
            self.serial_number = (highest_serial or 0) + 1
    
    def before_save(self):
        project_calculations.calculate_all_values(self)
    
    def get_party(self, party_name):
        """Get party details from project"""
        return next((p for p in self.parties if p.party == party_name), None)

    @frappe.whitelist()
    def add_scope(self, scope_data):
        """Add a new scope to the project"""
        scope_data = frappe.parse_json(scope_data)
        
        if not self.scopes:
            self.scopes = []
            
        next_scope_number = (
            max([s.scope_number for s in self.scopes]) + 1 if self.scopes else 1
        )
        
        scope_data['scope_number'] = next_scope_number
        self.append('scopes', scope_data)
        self.save()
        
        return self.scopes[-1]

#region Excel Import

def apply_style(cell, style):
    """Apply a predefined style to a cell"""
    # Apply each style property if it exists
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']


@frappe.whitelist()
def get_import_template(scope):
    scope = json.loads(scope)
    
    # Get project details
    project = frappe.get_doc('Project', scope.get('parent'))
    company_info = frappe.db.get_singles_dict('Rua')
    brand_color = company_info.get('brand_color', '#2C3E50').lstrip('#')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items Import"

    # Set column widths
    column_widths = {
        'A': 20,  # Item
        'B': 30,  # Description
        'C': 10,  # Qty
        'D': 12,  # Width
        'E': 12,  # Height
        'F': 15,  # Glass Unit
        'G': 15,  # Curtain Wall
        'H': 15,  # Insertion 1
        'I': 15,  # Insertion 2
        'J': 15,  # Insertion 3
        'K': 15,  # Insertion 4
        'L': 15,  # Profit Percentage
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    current_row = 1
    
    # Add logo if exists
    if company_info.get('logo_horizontal'):
        logo_path = frappe.get_site_path('public', company_info.logo_horizontal.lstrip('/'))
        if os.path.exists(logo_path):
            with PILImage.open(logo_path) as img:
                img = img.convert('RGB')
                max_height = 45  # pixels
                aspect_ratio = img.width / img.height
                new_width = int(max_height * aspect_ratio)
                img = img.resize((new_width, max_height))
                
                img_bytes = io.BytesIO()
                img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                logo = Image(img_bytes)
                logo.width = new_width
                logo.height = max_height
                ws.add_image(logo, 'A1')
        
        # Merge cells and set background color
        ws.merge_cells(f'C{current_row}:L{current_row}')
        for col in range(2, 13):  # C through L
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
                
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # Company Name - Large and prominent
    ws.row_dimensions[current_row].height = 35
    company_cell = ws.cell(row=current_row, column=1)
    company_cell.value = company_info.get('company_name', '')
    company_cell.font = Font(name='Calibri', size=18, bold=True, color=brand_color)
    company_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    # Project Serial Number and VAT info
    serial_cell = ws.cell(row=current_row, column=12)
    serial_cell.value = f"Serial No: {project.serial_number or 'Not Assigned'}  |  Scope: Scope {scope['scope_number']}{' - ' + scope['description'] if scope.get('description') else ''}"
    serial_cell.font = Font(name='Calibri', size=11, color='666666')
    serial_cell.alignment = Alignment(horizontal='right', vertical='center')
    current_row += 1

    # Separator line
    ws.row_dimensions[current_row].height = 4
    for col in range(1, 13):  # A through L
        cell = ws.cell(row=current_row, column=col)
        cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
    current_row += 1

    # Project Name and Location
    ws.row_dimensions[current_row].height = 30
    
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = project.project_name
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='2C3E50')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    location_cell = ws.cell(row=current_row, column=12)
    location_cell.value = f"Location: {project.location}"
    location_cell.font = Font(name='Calibri', size=11, color='666666')
    location_cell.alignment = Alignment(horizontal='right', vertical='center')
    current_row += 2

    # Headers
    headers = [
        'Item*', 'Description', 'Qty*', 'Width*', 'Height*', 'Glass Unit', 
        'Curtain Wall*', 'Insertion 1', 'Insertion 2', 'Insertion 3', 'Insertion 4', 
        'Profit Percentage'
    ]
    
    header_font = Font(name='Calibri', size=11, bold=True, color='2C3E50')
    header_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    header_border = Border(
        bottom=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add Instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions.merge_cells('A1:D1')
    title_cell = instructions.cell(row=1, column=1, value="Instructions & Notes")
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    
    notes = [
        ("Required Fields:", "Fields marked with * are mandatory"),
        ("Measurements:", "Width and Height should be in millimeters"),
        ("Glass Unit:", "Will use scope's default value if not provided"),
        ("Profit Percentage:", "Will use scope's default value if not provided"),
        ("Important:", "Only modify the data rows in the 'Items Import' sheet"),
        ("", "Do not modify headers or add rows above the headers")
    ]
    
    for i, (title, content) in enumerate(notes, 3):
        if title:
            title_cell = instructions.cell(row=i, column=1, value=title)
            title_cell.font = Font(name='Calibri', bold=True)
            title_cell.fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
        content_cell = instructions.cell(row=i, column=2, value=content)
        content_cell.alignment = Alignment(wrap_text=True)
    
    instructions.column_dimensions['A'].width = 20
    instructions.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    filename = f"RUA_{project.project_name}_{project.serial_number or 'NO_SERIAL'}_{scope['scope_number']}.xlsx"
    filename = re.sub(r'[^\w\-_.]', '_', filename)
    
    content = base64.b64encode(excel_file.getvalue()).decode('utf-8')
    
    return {
        'filename': filename,
        'content': content
    }
@frappe.whitelist()
def import_items_from_excel(file_url, scope):
    try:
        scope = json.loads(scope)
        
        # Get the file document
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = file_doc.get_full_path()
        
        # Read the Excel file
        wb = openpyxl.load_workbook(filename=file_path, data_only=True)
        ws = wb.active
        
        items = []
        # Start from row 7 (right after headers)
        for row in ws.iter_rows(min_row=7):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
                
            try:
                # Get glass unit and profit from scope if not provided in Excel
                excel_glass_unit = row[5].value
                excel_profit = row[11].value
                
                glass_unit = excel_glass_unit if excel_glass_unit not in [None, ''] else scope.get('glass_sqm_price')
                profit_percentage = float(excel_profit) if excel_profit not in [None, ''] else float(scope.get('profit', 0))
                
                item_data = {
                    'doctype': 'Project Items',
                    'parenttype': 'Project',
                    'parentfield': 'items',
                    'parent': scope.get('parent'),
                    'item': row[0].value,
                    'description': row[1].value or '',
                    'qty': float(row[2].value or 0),
                    'width': float(row[3].value or 0),
                    'height': float(row[4].value or 0),
                    'glass_unit': glass_unit,
                    'curtain_wall': row[6].value or 0,
                    'insertion_1': row[7].value or 0,
                    'insertion_2': row[8].value or 0,
                    'insertion_3': row[9].value or 0,
                    'insertion_4': row[10].value or 0,
                    'profit_percentage': profit_percentage,
                    'scope_number': scope.get('scope_number')
                }
                
                # Validate required fields
                if not all([item_data['item'], item_data['qty'], item_data['width'], item_data['height']]):
                    raise ValueError(f"Missing required fields in row {row[0].row}")
                    
                items.append(item_data)
            except (ValueError, TypeError) as e:
                frappe.throw(f"Error in row {row[0].row}: {str(e)}. Please check that all numeric fields contain valid numbers.")
        
        if not items:
            frappe.throw("No valid items found in the Excel file")
            
        # Get the project document
        project = frappe.get_doc('Project', scope.get('parent'))
        
        # Add items to the project
        for item in items:
            project.append('items', item)
        
        # Calculate values before saving to ensure new items are included
        project_calculations.calculate_all_values(project)
        
        # Save the project
        project.save()
        
        return {
            "message": f"Successfully imported {len(items)} items",
            "items": items
        }
    except Exception as e:
        frappe.throw(f"Error processing Excel file: {str(e)}")

#endregion

#region Document Generation

@frappe.whitelist()
def make_project_bill(source_name, target_doc=None):
    args = json.loads(frappe.form_dict.args)
    bill_type = args.get("bill_type")
    scope = args.get("scope", 1)  # Default to 1 if not specified
    send_rfq_link = args.get("send_rfq_link", 0)
    url = args.get("url")
    selected_items = args.get("selected_items", [])
    party = args.get("party")  # Get selected party
    
    source_doc = frappe.get_doc("Project", source_name)
    
    # Create a new Project Bill
    doclist = frappe.new_doc("Project Bill")
    doclist.project = source_doc.name
    doclist.serial_number = source_doc.serial_number
    doclist.date = frappe.utils.nowdate()
    doclist.bill_type = bill_type
    doclist.scope = scope
    doclist.vat = frappe.db.get_single_value('Rua', 'vat')
    
    # Set the selected party as the main party
    if party:
        doclist.party = party
    
    # Handle scope description
    if str(scope) == "0":
        # All Scopes case
        doclist.scope_description = f"Project {source_doc.name}: All Scopes"
    elif source_doc.scopes:
        # Single scope case
        scope_row = next((s for s in source_doc.scopes if str(s.scope_number) == str(scope)), None)
        if scope_row:
            doclist.scope_description = scope_row.description
    
    
    
    
    # Handle items based on bill type
    if bill_type == "Request for Quotation":
        if send_rfq_link:
            # RFQ from link case
            doclist.send_rfq_link = 1
            doclist.url = url
        elif selected_items:
            # RFQ from items case
            for item in selected_items:
                doclist.append("items", {
                    "item": item.get("item"),
                    "description": item.get("description"),
                    "qty": item.get("qty"),
                    "width": item.get("width"),
                    "height": item.get("height"),
                    "party": party
                })
    elif bill_type == "Purchase Order" and selected_items:
        # Purchase Order case
        for item in selected_items:
            doclist.append("items", {
                "item": item.get("item"),
                "description": item.get("description"),
                "qty": item.get("qty"),
                "width": item.get("width"),
                "height": item.get("height"),
                "rate": item.get("rate"),
                "party": party
            })
    elif bill_type == "Quotation" and source_doc.items:
        # Existing Quotation logic
        for item in source_doc.items:
            if str(scope) == "0" or str(item.scope_number) == str(scope):
                doclist.append("items", {
                    "item": item.item,
                    "description": item.description,
                    "qty": item.qty,
                    "width": item.width,
                    "height": item.height,
                    "rate": item.actual_unit_rate,
                    "party": party
                })
    elif bill_type in ["Proforma", "Tax Invoice"]:
        if str(scope) == "0" and len(source_doc.scopes) > 1:
            # All Scopes case - create an item for each scope
            for scope_row in source_doc.scopes:
                doclist.append("items", {
                    "item": source_doc.name,
                    "description": scope_row.description,
                    "qty": 1,
                    "rate": scope_row.total_price or 0,
                    "party": party
                })
        else:
            # Single scope case
            description = "Supply & Installation of Glass and Aluminum Works"
            if source_doc.scopes:
                scope_row = next((s for s in source_doc.scopes if str(s.scope_number) == str(scope)), None)
                if scope_row and scope_row.description:
                    description = scope_row.description
                    rate = scope_row.total_price or 0
                else:
                    rate = 0
            else:
                rate = 0
                
            doclist.append("items", {
                "item": source_doc.name,
                "description": description,
                "qty": 1,
                "rate": rate,
                "party": party
            })
    
    doclist.calculate_amounts()
    doclist.calculate_totals()
    return doclist

@frappe.whitelist()
def make_payment_voucher(source_name, target_doc=None):
    """Create a Payment Voucher from Project"""
    # Get args from form_dict
    args = frappe.form_dict.get('args')
    if args:
        try:
            args = json.loads(args)
            party = args.get('party')
            outstanding_amount = float(args.get('outstanding_amount', 0))
        except:
            party = None
            outstanding_amount = 0
    else:
        party = None
        outstanding_amount = 0

    if not party:
        frappe.throw("Party is required to create Payment Voucher")

    project = frappe.get_doc("Project", source_name)
    party_doc = project.get_party(party)
    if not party_doc:
        frappe.throw(f"Party {party} not found in project {source_name}")

    target = frappe.new_doc("Payment Voucher")
    target.project = source_name
    target.party = party
    target.type = "Pay" if party_doc.type == "Supplier" else "Receive"
    
    # Set amount based on outstanding
    if outstanding_amount != 0:
        target.amount = abs(outstanding_amount)  # Always use positive amount
        
    return target

@frappe.whitelist()
def get_party_outstanding_amounts(project, parties):
    """Get outstanding amounts for multiple parties"""
    parties = frappe.parse_json(parties)
    result = {}
    
    for party_data in parties:
        party = party_data.get('party')
        party_type = party_data.get('type')
        
        if party_type == 'Client':
            # Get total of submitted Tax Invoices
            total_billed = frappe.db.sql("""
                SELECT COALESCE(SUM(grand_total), 0) as total
                FROM `tabProject Bill`
                WHERE project = %s
                AND party = %s
                AND bill_type = 'Tax Invoice'
                AND docstatus = 1
            """, (project, party))[0][0]
            
            # Get total received payments
            total_paid = frappe.db.sql("""
                SELECT COALESCE(SUM(amount), 0) as total
                FROM `tabPayment Voucher`
                WHERE project = %s
                AND party = %s
                AND type = 'Receive'
                AND docstatus = 1
            """, (project, party))[0][0]
            
        else:  # Supplier
            # Get total of submitted Purchase Orders
            total_billed = frappe.db.sql("""
                SELECT COALESCE(SUM(grand_total), 0) as total
                FROM `tabProject Bill`
                WHERE project = %s
                AND party = %s
                AND bill_type = 'Purchase Order'
                AND docstatus = 1
            """, (project, party))[0][0]
            
            # Get total paid to supplier
            total_paid = frappe.db.sql("""
                SELECT COALESCE(SUM(amount), 0) as total
                FROM `tabPayment Voucher`
                WHERE project = %s
                AND party = %s
                AND type = 'Pay'
                AND docstatus = 1
            """, (project, party))[0][0]
        
        result[party] = total_billed - total_paid
    
    return result

#endregion

@frappe.whitelist()
def add_item(project, scope_number, item_data):
    """Add an item to a project scope and calculate all values
    
    Args:
        project (str): Project name
        scope_number (str): Scope number
        item_data (dict): Item data including all required fields
    """
    doc = frappe.get_doc("Project", project)
    scope = next((s for s in doc.scopes if str(s.scope_number) == str(scope_number)), None)
    
    if not scope:
        frappe.throw(f"Scope {scope_number} not found in project {project}")
        
    # Convert item_data from string to dict if needed
    if isinstance(item_data, str):
        item_data = frappe.parse_json(item_data)
    
    # Ensure numeric fields are properly converted
    numeric_fields = ['width', 'height', 'glass_unit', 'curtain_wall', 
                     'insertion_1', 'insertion_2', 'insertion_3', 'insertion_4', 'qty', 'area']
    
    for field in numeric_fields:
        if field in item_data:
            item_data[field] = flt(item_data[field])
    
    # Handle manual area flag
    if item_data.get('manual_area'):
        item_data['width'] = 0
        item_data['height'] = 0
    
    # Create new item
    new_item = {
        "scope_number": scope_number,
        **item_data
    }
    
    # Add item to child table
    doc.append("items", new_item)
    
    # First save to add the item
    doc.save()
    
    # Get a fresh copy of the doc with the new item
    updated_doc = frappe.get_doc("Project", project)
    
    # Force recalculation
    project_calculations.calculate_all_values(updated_doc)
    
    # Save again to persist calculations
    updated_doc.save(ignore_permissions=True)
    
    return new_item
    
@frappe.whitelist()
def get_item_suggestions(doctype, txt, searchfield, start, page_len, filters):
    """Get item suggestions for autocomplete with party history and dimensions"""
    try:
        filters = json.loads(filters) if isinstance(filters, str) else filters
        party = filters.get('party') if filters else None
        
        # Base conditions and values
        conditions = []
        values = {
            'start': int(start),
            'page_len': int(page_len)
        }
        
        if txt:
            conditions.append("i.item LIKE %(txt)s")
            values['txt'] = f'%{txt}%'
        
        # If party is selected, get their items first with most recent values
        order_by = "i.item"
        if party:
            values['party'] = party
            query = f"""
                WITH RankedItems AS (
                    SELECT 
                        i.item,
                        i.description,
                        i.rate,
                        i.width,
                        i.height,
                        ROW_NUMBER() OVER (
                            PARTITION BY i.item 
                            ORDER BY i.modified DESC
                        ) as rn
                    FROM `tabItems` i
                    WHERE i.party = %(party)s
                    {' AND ' + ' AND '.join(conditions) if conditions else ''}
                )
                SELECT 
                    item,
                    description,
                    COALESCE(rate, 0) as last_rate,
                    COALESCE(width, 0) as width,
                    COALESCE(height, 0) as height
                FROM RankedItems
                WHERE rn = 1
                ORDER BY item
                LIMIT %(start)s, %(page_len)s
            """
        else:
            # For non-party searches, show distinct items with their most recent values
            query = f"""
                WITH RankedItems AS (
                    SELECT 
                        i.item,
                        i.description,
                        i.rate,
                        i.width,
                        i.height,
                        ROW_NUMBER() OVER (
                            PARTITION BY i.item 
                            ORDER BY i.modified DESC
                        ) as rn
                    FROM `tabItems` i
                    {' WHERE ' + ' AND '.join(conditions) if conditions else ''}
                )
                SELECT 
                    item,
                    description,
                    COALESCE(rate, 0) as last_rate,
                    COALESCE(width, 0) as width,
                    COALESCE(height, 0) as height
                FROM RankedItems
                WHERE rn = 1
                ORDER BY item
                LIMIT %(start)s, %(page_len)s
            """
        
        result = frappe.db.sql(query, values, as_list=1)
        return result
        
    except Exception as e:
        frappe.logger("items").error(f"Item Suggestion Error: {str(e)}\nFilters: {filters}")
        return []