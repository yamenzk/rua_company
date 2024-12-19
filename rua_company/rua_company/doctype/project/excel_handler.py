import frappe
from frappe.utils import cstr, flt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.drawing.image import Image
from frappe.utils import get_site_base_path
import math
import json
import os
from PIL import Image as PILImage
import io
from . import project_calculations

def define_styles(brand_color):
    """Define consistent styles for the workbook"""
    return {
        'header': {
            'font': Font(name='Calibri', size=12, bold=True, color='171717'),
            'fill': PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
        },
        'input': {
            'font': Font(name='Calibri', size=10),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin', color='E5E5E5'),
                right=Side(style='thin', color='E5E5E5'),
                top=Side(style='thin', color='E5E5E5'),
                bottom=Side(style='thin', color='E5E5E5')
            )
        },
        'computed': {
            'font': Font(name='Calibri', size=10),
            'fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin', color='E5E5E5'),
                right=Side(style='thin', color='E5E5E5'),
                top=Side(style='thin', color='E5E5E5'),
                bottom=Side(style='thin', color='E5E5E5')
            )
        },
        'settings': {
            'font': Font(name='Calibri', size=11, bold=True, color='171717'),
            'fill': PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='B3B3B3'),
                right=Side(style='thin', color='B3B3B3'),
                top=Side(style='thin', color='B3B3B3'),
                bottom=Side(style='thin', color='B3B3B3')
            )
        }
    }

def apply_modern_branding(ws, company_info, scope, project_name):
    """Apply clean, modern branding with company logo and information"""
    brand_color = company_info.get('brand_color', '#2C3E50').lstrip('#')
    current_row = 1
    
    # Company logo
    if company_info.get('logo_horizontal'):
        logo_path = frappe.get_site_path('public', company_info.logo_horizontal.lstrip('/'))
        if os.path.exists(logo_path):
            with PILImage.open(logo_path) as img:
                img = img.convert('RGB')
                max_height = 45
                aspect_ratio = img.width / img.height
                new_width = int(max_height * aspect_ratio)
                img = img.resize((new_width, max_height))
                
                img_bytes = io.BytesIO()
                img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                logo = Image(img_bytes)
                logo.width = new_width
                logo.height = max_height
                ws.add_image(logo, 'A1')
                
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # Company name
    ws.row_dimensions[current_row].height = 35
    company_cell = ws.cell(row=current_row, column=1)
    company_cell.value = company_info.get('company_name', '')
    company_cell.font = Font(name='Calibri', size=18, bold=True, color=brand_color)
    company_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    company_cell.protection = Protection(locked=True)
    current_row += 1
    

    # Separator line
    ws.row_dimensions[current_row].height = 4
    for col in range(1, 27):  # Extend line across sheet
        cell = ws.cell(row=current_row, column=col)
        cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
    current_row += 1

    # Report title
    ws.row_dimensions[current_row].height = 30
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Scope {scope.scope_number}: {scope.description} - Project: {project_name}"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='2C3E50')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    title_cell.protection = Protection(locked=True)
    current_row += 2

    return current_row

def apply_style(cell, style):
    """Apply a predefined style to a cell"""
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']

@frappe.whitelist()
def get_items_template(project_name):
    project = frappe.get_doc("Project", project_name)
    wb = openpyxl.Workbook()
    company_info = frappe.db.get_singles_dict('Rua')
    brand_color = company_info.get('brand_color', '#2C3E50').lstrip('#')
    styles = define_styles(brand_color)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Group items by scope
    items_by_scope = {}
    for item in project.items:
        items_by_scope.setdefault(str(item.scope_number), []).append(item)

    for scope in project.scopes:
        scope_name = f"Scope{scope.scope_number}"
        ws = wb.create_sheet(title=scope_name)
        
        current_row = apply_modern_branding(ws, company_info, scope, project_name)
        
        # Settings section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        header_cell = ws.cell(row=current_row, column=1, value="Scope Settings")
        apply_style(header_cell, styles['header'])
        current_row += 1
        
        settings = [
            ("VAT", flt(scope.vat, 2)),
            ("LabourCharges", flt(scope.labour_charges, 2)),
            ("DefaultProfit", flt(scope.profit, 2)),
            ("Rounding", scope.rounding or "None")
        ]
        
        if scope.type in ["Openings", "Skylights"]:
            settings.extend([
                ("GlassSQMPrice", flt(scope.glass_sqm_price, 2)),
                ("AluminumWeight", flt(scope.aluminum_weight, 2)),
                ("SDF", flt(scope.sdf, 2))
            ])
            
        # Add settings with modern styling
        for key, value in settings:
            label_cell = ws.cell(row=current_row, column=1, value=key.replace("_", " "))
            value_cell = ws.cell(row=current_row, column=2, value=value)
            
            for cell in [label_cell, value_cell]:
                apply_style(cell, styles['settings'])
                cell.protection = Protection(locked=True)
            
            range_name = f'{scope_name}_{key}'
            defined_name = DefinedName(name=range_name, attr_text=f"'{scope_name}'!$B${current_row}")
            wb.defined_names[range_name] = defined_name
            current_row += 1

        # Add computed ratio for Openings/Skylights
        if scope.type in ["Openings", "Skylights"]:
            ratio_cell = ws.cell(row=current_row, column=1, value="Ratio")
            apply_style(ratio_cell, styles['settings'])
            
            formula_cell = ws.cell(row=current_row, column=2)
            apply_style(formula_cell, styles['computed'])
            
            range_name = f'{scope_name}_Ratio'
            defined_name = DefinedName(name=range_name, attr_text=f"'{scope_name}'!$B${current_row}")
            wb.defined_names[range_name] = defined_name
            current_row += 1

        current_row += 1

        # Define columns based on scope type
        if scope.type in ["Openings", "Skylights"]:
            input_cols = [
                "Item", "Description", "Quantity", "Width (cm)", "Height (cm)",
                "Manual Area", "Glass Unit", "Curtain Wall",
                "Insertion 1", "Insertion 2", "Insertion 3", "Insertion 4",
                "Profit %"
            ]
            computed_cols = [
                "Area", "Glass Price", "Total Glass",
                "Aluminum Price", "Aluminum Unit", "Total Aluminum",
                "Actual Unit", "Total Profit", "Total Cost",
                "Actual Unit Rate", "Overall Price"
            ]
        else:
            input_cols = [
                "Item", "Description", "Quantity",
                "Actual Unit Rate", "Profit %"
            ]
            if scope.type not in ["Handrails", "Cladding"]:
                input_cols.insert(3, "Width (cm)")
                input_cols.insert(4, "Height (cm)")
                input_cols.insert(5, "Manual Area")
                input_cols.insert(6, "Area")
            
            computed_cols = [
                "Actual Unit", "Total Profit",
                "Total Cost", "Overall Price"
            ]

        # Write headers and prepare column mapping
        header_row = current_row
        col_mapping = {}
        
        # Write input columns
        for idx, header in enumerate(input_cols, 1):
            cell = ws.cell(row=header_row, column=idx, value=header)
            apply_style(cell, styles['header'])
            cell.protection = Protection(locked=True)
            col_mapping[header] = get_column_letter(idx)
            if idx == 2:  # Column B
                ws.column_dimensions[get_column_letter(idx)].width = 40  # or whatever width you prefer
            else:
                ws.column_dimensions[get_column_letter(idx)].width = 15


        # Write computed columns
        for idx, header in enumerate(computed_cols, len(input_cols) + 1):
            cell = ws.cell(row=header_row, column=idx, value=header)
            apply_style(cell, styles['header'])
            col_mapping[header] = get_column_letter(idx)
            ws.column_dimensions[get_column_letter(idx)].width = 15
            cell.protection = Protection(locked=True)

        # Add formulas
        data_start_row = header_row + 1
        
        if scope.type in ["Openings", "Skylights"]:
            start_row = data_start_row
            end_row = data_start_row + 99
            
            # Get column letters for aluminum components
            curtain_col = col_mapping["Curtain Wall"]
            ins1_col = col_mapping["Insertion 1"]
            ins2_col = col_mapping["Insertion 2"]
            ins3_col = col_mapping["Insertion 3"]
            ins4_col = col_mapping["Insertion 4"]
            qty_col = col_mapping["Quantity"]

            # Build the aluminum sum array formula
            aluminum_sum = (
                f'SUMPRODUCT('
                f'{curtain_col}{start_row}:{curtain_col}{end_row}+'
                f'{ins1_col}{start_row}:{ins1_col}{end_row}+'
                f'{ins2_col}{start_row}:{ins2_col}{end_row}+'
                f'{ins3_col}{start_row}:{ins3_col}{end_row}+'
                f'{ins4_col}{start_row}:{ins4_col}{end_row},'
                f'{qty_col}{start_row}:{qty_col}{end_row})'
            )

            # Complete ratio formula
            ratio_formula = (
                f'=IF({aluminum_sum}=0,1,'
                f'({scope_name}_AluminumWeight*{scope_name}_SDF+{aluminum_sum}*'
                f'(1+{scope_name}_VAT/100))/{aluminum_sum})'
            )
            formula_cell.value = ratio_formula

            # Item row formulas
            formulas = {
                "Area": (
                    f'=IF({col_mapping["Manual Area"]}{{row}}<>0,'
                    f'{col_mapping["Manual Area"]}{{row}},'
                    f'({col_mapping["Width (cm)"]}{{row}}*{col_mapping["Height (cm)"]}{{row}})/10000)'
                ),
                "Glass Price": (
                    f'=IF({col_mapping["Glass Unit"]}{{row}}>=0,'
                    f'{col_mapping["Glass Unit"]}{{row}}*{col_mapping["Area"]}{{row}}*(1+{scope_name}_VAT/100),0)'
                ),
                "Total Glass": (
                    f'={col_mapping["Glass Price"]}{{row}}*{col_mapping["Quantity"]}{{row}}'
                ),
                "Aluminum Price": (
                    f'={col_mapping["Curtain Wall"]}{{row}}+'
                    f'{col_mapping["Insertion 1"]}{{row}}+'
                    f'{col_mapping["Insertion 2"]}{{row}}+'
                    f'{col_mapping["Insertion 3"]}{{row}}+'
                    f'{col_mapping["Insertion 4"]}{{row}}'
                ),
                "Aluminum Unit": (
                    f'={col_mapping["Aluminum Price"]}{{row}}*{scope_name}_Ratio'
                ),
                "Total Aluminum": (
                    f'={col_mapping["Aluminum Unit"]}{{row}}*{col_mapping["Quantity"]}{{row}}'
                ),
                "Actual Unit": (
                    f'={col_mapping["Aluminum Unit"]}{{row}}+{col_mapping["Glass Price"]}{{row}}+'
                    f'{scope_name}_LabourCharges'
                ),
                "Total Profit": (
                    f'={col_mapping["Actual Unit"]}{{row}}*{col_mapping["Profit %"]}{{row}}/100'
                ),
                "Total Cost": (
                    f'={col_mapping["Actual Unit"]}{{row}}*{col_mapping["Quantity"]}{{row}}'
                ),
                "Actual Unit Rate": (
                    f'={col_mapping["Total Profit"]}{{row}}+{col_mapping["Actual Unit"]}{{row}}'
                ),
                "Overall Price": (
                    f'=IF({scope_name}_Rounding="Round up to nearest 5",'
                    f'CEILING({col_mapping["Actual Unit Rate"]}{{row}}*{col_mapping["Quantity"]}{{row}}/5)*5,'
                    f'{col_mapping["Actual Unit Rate"]}{{row}}*{col_mapping["Quantity"]}{{row}})'
                )
            }
        else:
            formulas = {
                "Actual Unit": (
                    f'={col_mapping["Actual Unit Rate"]}{{row}}-'
                    f'({col_mapping["Actual Unit Rate"]}{{row}}*{col_mapping["Profit %"]}{{row}}/100)+'
                    f'{scope_name}_LabourCharges'
                ),
                "Total Profit": (
                    f'={col_mapping["Actual Unit Rate"]}{{row}}*{col_mapping["Profit %"]}{{row}}/100'
                ),
                "Total Cost": (
                    f'={col_mapping["Actual Unit"]}{{row}}*{col_mapping["Quantity"]}{{row}}'
                )
            }
            
            if scope.type in ["Handrails", "Cladding"]:
                formulas["Overall Price"] = (
                    f'=IF({scope_name}_Rounding="Round up to nearest 5",'
                    f'CEILING({col_mapping["Quantity"]}{{row}}*{col_mapping["Actual Unit Rate"]}{{row}}/5)*5,'
                    f'{col_mapping["Quantity"]}{{row}}*{col_mapping["Actual Unit Rate"]}{{row}})'
                )
            else:
                formulas["Area"] = (
                    f'=IF({col_mapping["Manual Area"]}{{row}}<>0,'
                    f'{col_mapping["Manual Area"]}{{row}},'
                    f'({col_mapping["Width (cm)"]}{{row}}*{col_mapping["Height (cm)"]}{{row}})/10000)'
                )
                formulas["Overall Price"] = (
                    f'=IF({scope_name}_Rounding="Round up to nearest 5",'
                    f'CEILING({col_mapping["Quantity"]}{{row}}*{col_mapping["Area"]}{{row}}*'
                    f'{col_mapping["Actual Unit Rate"]}{{row}}/5)*5,'
                    f'{col_mapping["Quantity"]}{{row}}*{col_mapping["Area"]}{{row}}*'
                    f'{col_mapping["Actual Unit Rate"]}{{row}})'
                )

        # Apply formulas to rows
        for row in range(data_start_row, data_start_row + 100):  # Prepare 100 rows
            # Format input columns
            for header in input_cols:
                col = col_mapping[header]
                cell = ws.cell(row=row, column=ws[col + '1'].column)
                apply_style(cell, styles['input'])
                cell.protection = Protection(locked=False)
            
            # Apply formulas to computed columns
            for header, formula in formulas.items():
                col = col_mapping[header]
                cell = ws.cell(row=row, column=ws[col + '1'].column)
                cell.value = formula.format(row=row)
                apply_style(cell, styles['computed'])
                cell.protection = Protection(locked=True)
                # Apply AED format to all computed columns except Area
                if header != "Area":
                    cell.number_format = '#,##0.00 "AED"'
                else:
                    cell.number_format = '#,##0.00'  # Just decimal format for Area

        # Fill existing data
        if str(scope.scope_number) in items_by_scope:
            for idx, item in enumerate(items_by_scope[str(scope.scope_number)]):
                row = data_start_row + idx
                
                # Base fields for all types
                ws.cell(row=row, column=ws[col_mapping["Item"] + '1'].column).value = item.item
                ws.cell(row=row, column=ws[col_mapping["Description"] + '1'].column).value = item.description
                ws.cell(row=row, column=ws[col_mapping["Quantity"] + '1'].column).value = item.qty
                ws.cell(row=row, column=ws[col_mapping["Profit %"] + '1'].column).value = item.profit_percentage
                
                if scope.type in ["Openings", "Skylights"]:
                    if item.manual_area:
                        ws.cell(row=row, column=ws[col_mapping["Width (cm)"] + '1'].column).value = 0
                        ws.cell(row=row, column=ws[col_mapping["Height (cm)"] + '1'].column).value = 0
                        ws.cell(row=row, column=ws[col_mapping["Manual Area"] + '1'].column).value = item.area
                    else:
                        ws.cell(row=row, column=ws[col_mapping["Width (cm)"] + '1'].column).value = item.width
                        ws.cell(row=row, column=ws[col_mapping["Height (cm)"] + '1'].column).value = item.height
                        ws.cell(row=row, column=ws[col_mapping["Manual Area"] + '1'].column).value = 0
                        
                    ws.cell(row=row, column=ws[col_mapping["Glass Unit"] + '1'].column).value = item.glass_unit
                    ws.cell(row=row, column=ws[col_mapping["Curtain Wall"] + '1'].column).value = item.curtain_wall
                    ws.cell(row=row, column=ws[col_mapping["Insertion 1"] + '1'].column).value = item.insertion_1
                    ws.cell(row=row, column=ws[col_mapping["Insertion 2"] + '1'].column).value = item.insertion_2
                    ws.cell(row=row, column=ws[col_mapping["Insertion 3"] + '1'].column).value = item.insertion_3
                    ws.cell(row=row, column=ws[col_mapping["Insertion 4"] + '1'].column).value = item.insertion_4
                else:
                    ws.cell(row=row, column=ws[col_mapping["Actual Unit Rate"] + '1'].column).value = item.actual_unit_rate
                    if scope.type not in ["Handrails", "Cladding"]:
                        if item.manual_area:
                            ws.cell(row=row, column=ws[col_mapping["Width (cm)"] + '1'].column).value = 0
                            ws.cell(row=row, column=ws[col_mapping["Height (cm)"] + '1'].column).value = 0
                            ws.cell(row=row, column=ws[col_mapping["Manual Area"] + '1'].column).value = item.area
                        else:
                            ws.cell(row=row, column=ws[col_mapping["Width (cm)"] + '1'].column).value = item.width
                            ws.cell(row=row, column=ws[col_mapping["Height (cm)"] + '1'].column).value = item.height
                            ws.cell(row=row, column=ws[col_mapping["Manual Area"] + '1'].column).value = 0

        # Protect the worksheet
        ws.protection.sheet = True
        
        # Freeze panes
        ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    return save_workbook_and_get_url(wb, project_name)

def save_workbook_and_get_url(wb, project_name):
    file_name = f"project_items_template_{project_name}.xlsx"
    file_path = frappe.get_site_path('private', 'files', file_name)
    
    wb.save(file_path)
    
    # Create a File document
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "folder": "Home/Attachments",
        "is_private": 1,
        "file_url": f"/private/files/{file_name}"
    })
    file_doc.insert(ignore_permissions=True)
    
    return file_doc.file_url

@frappe.whitelist()
def import_items_from_excel(project_name, file_url):
    project = frappe.get_doc("Project", project_name)
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Clear existing items
    project.items = []
    
    items_to_add = []
    for sheet in wb.worksheets:
        scope_number = sheet.title.replace("Scope", "")  # Updated to match new sheet naming
        scope = next((s for s in project.scopes if str(s.scope_number) == scope_number), None)
        if not scope:
            continue
            
        # Find the start of data rows
        data_start_row = None
        header_row = None
        for row_idx in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == "Item":
                header_row = row_idx
                data_start_row = row_idx + 1
                break
                
        if not data_start_row:
            continue

        # Map column headers to indices
        headers = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=header_row, column=col).value
            if header:
                headers[header] = col

        # Process data rows
        for row in range(data_start_row, sheet.max_row + 1):
            item_name = sheet.cell(row=row, column=headers["Item"]).value
            if not item_name:  # Skip empty rows
                continue

            item_dict = {
                "doctype": "Project Item",
                "scope_number": scope_number,
                "parenttype": "Project",
                "parentfield": "items",
                "parent": project_name,
                "item": item_name,
                "description": sheet.cell(row=row, column=headers["Description"]).value or "",
                "qty": flt(sheet.cell(row=row, column=headers["Quantity"]).value or 0),
                "profit_percentage": flt(sheet.cell(row=row, column=headers["Profit %"]).value or scope.profit or 0)
            }

            # Add type-specific fields
            if scope.type in ["Openings", "Skylights"]:
                manual_area_value = flt(sheet.cell(row=row, column=headers["Manual Area"]).value or 0)
                
                # If manual_area has a value greater than 0
                if manual_area_value > 0:
                    item_dict.update({
                        "width": 0,
                        "height": 0,
                        "manual_area": 1,  # Set to 1 to indicate manual area is being used
                        "area": manual_area_value,  # Use the manual_area value as the area
                        "glass_unit": flt(sheet.cell(row=row, column=headers["Glass Unit"]).value or scope.glass_sqm_price or 0),
                        "curtain_wall": flt(sheet.cell(row=row, column=headers["Curtain Wall"]).value or 0),
                        "insertion_1": flt(sheet.cell(row=row, column=headers["Insertion 1"]).value or 0),
                        "insertion_2": flt(sheet.cell(row=row, column=headers["Insertion 2"]).value or 0),
                        "insertion_3": flt(sheet.cell(row=row, column=headers["Insertion 3"]).value or 0),
                        "insertion_4": flt(sheet.cell(row=row, column=headers["Insertion 4"]).value or 0)
                    })
                else:
                    # Normal case when manual_area is 0
                    item_dict.update({
                        "width": flt(sheet.cell(row=row, column=headers["Width (cm)"]).value or 0),
                        "height": flt(sheet.cell(row=row, column=headers["Height (cm)"]).value or 0),
                        "manual_area": 0,
                        "glass_unit": flt(sheet.cell(row=row, column=headers["Glass Unit"]).value or 0),
                        "curtain_wall": flt(sheet.cell(row=row, column=headers["Curtain Wall"]).value or 0),
                        "insertion_1": flt(sheet.cell(row=row, column=headers["Insertion 1"]).value or 0),
                        "insertion_2": flt(sheet.cell(row=row, column=headers["Insertion 2"]).value or 0),
                        "insertion_3": flt(sheet.cell(row=row, column=headers["Insertion 3"]).value or 0),
                        "insertion_4": flt(sheet.cell(row=row, column=headers["Insertion 4"]).value or 0)
                    })
            else:
                item_dict["actual_unit_rate"] = flt(sheet.cell(row=row, column=headers["Actual Unit Rate"]).value or 0)
                if scope.type not in ["Handrails", "Cladding"]:
                    manual_area_value = flt(sheet.cell(row=row, column=headers["Manual Area"]).value or 0)
                    
                    if manual_area_value > 0:
                        item_dict.update({
                            "width": 0,
                            "height": 0,
                            "manual_area": 1,
                            "area": manual_area_value
                        })
                    else:
                        item_dict.update({
                            "width": flt(sheet.cell(row=row, column=headers["Width (cm)"]).value or 0),
                            "height": flt(sheet.cell(row=row, column=headers["Height (cm)"]).value or 0),
                            "manual_area": 0
                        })

            items_to_add.append(item_dict)
    
    if items_to_add:
        project.extend("items", items_to_add)
        project.save()
        updated_doc = frappe.get_doc("Project", project)
        project_calculations.calculate_all_values(updated_doc)
        updated_doc.save(ignore_permissions=True)
        
    return {"message": f"Successfully imported {len(items_to_add)} items"}