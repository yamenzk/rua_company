# Copyright (c) 2024, Yamen Zakhour and contributors
# For license information, please see license.txt

import frappe
import json
from frappe.model.document import Document
from frappe.utils import flt, cint
import math
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import os
from frappe.utils.file_manager import save_file


class ScopeItems(Document):
    def validate(self):
        self.load_custom_functions()
        self.calculate_item_values()
        self.calculate_totals()

    def load_custom_functions(self):
        """Load custom functions for calculations"""
        self.custom_functions = {}
        
        functions = frappe.get_all(
            "Scope Custom Function",
            fields=["function_name", "function_code", "parameters", "disabled"],
            filters={"disabled": 0}  # Only load enabled functions
        )
        
        for func in functions:
            try:
                params = [p.strip() for p in (func.parameters or "").split('\n') if p.strip()]
                param_str = ', '.join(params)
                
                # Create the function code with result variable pattern
                full_code = f"def {func.function_name}({param_str}):\n"
                full_code += '    result = None\n'  # Initialize result variable
                full_code += '\n'.join(f"    {line}" for line in func.function_code.split('\n'))
                full_code += "\n    if result is None:\n"
                full_code += "        frappe.throw('Function must set the result variable')\n"
                full_code += "    return result\n"
                
                # Create a local namespace with required modules
                local_ns = {}
                
                # Execute the function definition in the local namespace
                exec(full_code, {"frappe": frappe, "math": math, "flt": flt, "cint": cint}, local_ns)
                
                # Store the function in our custom functions dict
                self.custom_functions[func.function_name] = local_ns[func.function_name]
                
                # Test the function
                try:
                    test_args = [1.0] * len(params)  # Test with 1.0 for each parameter
                    self.custom_functions[func.function_name](*test_args)
                except Exception as e:
                    frappe.throw(f"Test failed for {func.function_name}: {str(e)}")
                    
            except Exception as e:
                frappe.log_error(
                    f"Error loading custom function {func.function_name}: {str(e)}"
                )

    def get_eval_context(self, variables, doc_totals):
        """Get evaluation context for formulas"""
        class CustomFunctions:
            def __init__(self, functions):
                for name, func in functions.items():
                    setattr(self, name, func)
        
        # Get constants from _constants_data
        constants = {}
        if hasattr(self, '_constants_data') and self._constants_data:
            try:
                constants = json.loads(self._constants_data)
            except Exception as e:
                frappe.log_error(f"Error loading constants data: {str(e)}")
        
        # Convert all constant values to float for calculations
        constants = {k: flt(v) for k, v in constants.items()}
        
        context = {
            "variables": variables,
            "doc_totals": doc_totals,
            "constants": constants,
            "custom": CustomFunctions(self.custom_functions),
            "frappe": frappe,
            "math": math,
            "flt": flt,
            "cint": cint
        }
        
        # Add constants directly to the context for backward compatibility
        context.update(constants)
        
        return context

    def calculate_item_values(self):
        """Calculate values for each item based on scope type formulas"""
        if not self.scope_type or not self.items:
            return

        scope_type = frappe.get_doc("Scope Type", self.scope_type)
        doc_totals = self.get_scope_totals()

        # Load custom functions if not already loaded
        if not hasattr(self, 'custom_functions'):
            self.load_custom_functions()

        # Sort fields by dependencies
        def get_dependencies(formula):
            if not formula:
                return set()
            deps = set()
            for field_name in [f.field_name for f in scope_type.scope_fields]:
                if f"variables['{field_name}']" in formula:
                    deps.add(field_name)
            # Check for doc_totals dependencies
            if formula and "doc_totals[" in formula:
                deps.add("__doc_totals__")  # Special marker for doc_totals dependency
            return deps

        # Create dependency graph
        field_deps = {}
        doc_totals_dependent_fields = []
        regular_fields = []
        
        for field in scope_type.scope_fields:
            if field.auto_calculate and field.calculation_formula:
                deps = get_dependencies(field.calculation_formula)
                if "__doc_totals__" in deps:
                    doc_totals_dependent_fields.append(field)
                    deps.remove("__doc_totals__")
                else:
                    regular_fields.append(field)
                field_deps[field.field_name] = deps

        # Sort regular fields by dependencies
        calculated_fields = set()
        sorted_regular_fields = []
        
        def add_field_to_list(field, target_list):
            if field.field_name in calculated_fields:
                return
            if field.auto_calculate and field.calculation_formula:
                deps = field_deps.get(field.field_name, set())
                for dep_name in deps:
                    dep_field = next((f for f in scope_type.scope_fields if f.field_name == dep_name), None)
                    if dep_field:
                        # Only add dependency if it's in the same category (regular or doc_totals)
                        if (dep_field in regular_fields and target_list == sorted_regular_fields) or \
                           (dep_field in doc_totals_dependent_fields and target_list == sorted_doc_totals_fields):
                            add_field_to_list(dep_field, target_list)
            calculated_fields.add(field.field_name)
            if field not in target_list:  # Prevent duplicates
                target_list.append(field)

        # Sort non-doc_totals dependent fields
        for field in regular_fields:
            add_field_to_list(field, sorted_regular_fields)

        # First calculate fields that don't depend on doc_totals
        for item in self.items:
            variables = self.get_item_variables(item)
            for field in sorted_regular_fields:
                if field.auto_calculate and field.calculation_formula:
                    try:
                        eval_globals = self.get_eval_context(variables, doc_totals)
                        result = eval(field.calculation_formula, eval_globals)
                        
                        if field.field_type == 'Int':
                            result = cint(result)
                        else:  # Float, Currency, or Percent
                            result = flt(result)
                            
                        item.set_dynamic_value(field.field_name, result)
                        variables[field.field_name] = result
                    except Exception as e:
                        frappe.throw(f"Error calculating {field.label}: {str(e)}")

        # Sort doc_totals dependent fields by their dependencies (excluding doc_totals)
        sorted_doc_totals_fields = []
        calculated_fields = set()  # Reset calculated fields for doc_totals sorting
        
        for field in doc_totals_dependent_fields:
            add_field_to_list(field, sorted_doc_totals_fields)

        # Now handle doc_totals dependent fields with multiple passes
        MAX_PASSES = 10  # Prevent infinite loops
        for pass_num in range(MAX_PASSES):
            # Store previous values to check for convergence
            previous_values = {}
            for item in self.items:
                previous_values[item.name] = {
                    field.field_name: item.get_dynamic_value(field.field_name)
                    for field in doc_totals_dependent_fields
                }
            
            # Calculate totals
            self.calculate_totals()
            doc_totals = self.get_scope_totals()
            
            # Calculate doc_totals dependent fields
            max_diff = 0
            for item in self.items:
                variables = self.get_item_variables(item)
                # Update variables with current field values
                for field in scope_type.scope_fields:
                    value = item.get_dynamic_value(field.field_name)
                    if value is not None:
                        variables[field.field_name] = value
                        
                for field in sorted_doc_totals_fields:
                    try:
                        eval_globals = self.get_eval_context(variables, doc_totals)
                        result = eval(field.calculation_formula, eval_globals)
                        
                        if field.field_type == 'Int':
                            result = cint(result)
                        else:  # Float, Currency, or Percent
                            result = flt(result)
                        
                        # Calculate maximum difference from previous value
                        prev_value = previous_values[item.name].get(field.field_name, 0)
                        max_diff = max(max_diff, abs(flt(result) - flt(prev_value)))
                        
                        item.set_dynamic_value(field.field_name, result)
                        variables[field.field_name] = result
                    except Exception as e:
                        frappe.throw(f"Error calculating {field.label}: {str(e)}")
            
            # Check if values have stabilized (convergence)
            if max_diff < 0.0001:  # Tolerance for floating point differences
                break
                
            if pass_num == MAX_PASSES - 1:
                frappe.msgprint(
                    "Warning: Calculations did not fully converge after maximum passes. "
                    "Some values might be approximate.",
                    indicator='orange'
                )

        # Final totals calculation
        self.calculate_totals()

    def calculate_totals(self):
        """Calculate scope-level totals"""
        if not self.scope_type or not self.items:
            # Clear totals if no items
            self._totals_data = json.dumps({})
            return

        scope_type = frappe.get_doc("Scope Type", self.scope_type)
        totals = {}  # Initialize totals dictionary
        
        # Get all items with their variables
        items_data = [self.get_item_variables(item) for item in self.items]
        
        def parse_filter_condition(condition):
            """Parse a single filter condition into field, operator, and value"""
            # Support for different operators
            operators = {
                '===': '==',
                '!==': '!=',
                '>=': '>=',
                '<=': '<=',
                '>': '>',
                '<': '<'
            }
            
            for js_op, py_op in operators.items():
                if js_op in condition:
                    field = condition.split('item.')[1].split(' ')[0]
                    value = condition.split(js_op)[1].split(')')[0].strip()
                    
                    # Convert value to appropriate type
                    if value.isdigit():
                        value = int(value)
                    elif value.replace('.', '').isdigit():
                        value = float(value)
                    elif value.lower() in ('true', 'false'):
                        value = value.lower() == 'true'
                    elif value.startswith('"') or value.startswith("'"):
                        value = value.strip("'\"")
                        
                    return field, py_op, value
            
            return None, None, None
            
        def evaluate_condition(item, field, op, value):
            """Evaluate a single condition against an item"""
            item_value = item.get(field)
            if item_value is None:
                return False
                
            if op == '==':
                return item_value == value
            elif op == '!=':
                return item_value != value
            elif op == '>=':
                return item_value >= value
            elif op == '<=':
                return item_value <= value
            elif op == '>':
                return item_value > value
            elif op == '<':
                return item_value < value
            
            return False

        # Calculate each formula
        for formula in scope_type.calculation_formulas:
            try:
                # Detect if this is a filtered aggregate formula
                if 'items.filter' in formula.formula:
                    filter_part = formula.formula.split('items.filter(')[1].split('.reduce')[0]
                    
                    # Handle multiple conditions with && or ||
                    conditions = []
                    if '&&' in filter_part:
                        condition_parts = filter_part.split('&&')
                        combine_op = all
                    elif '||' in filter_part:
                        condition_parts = filter_part.split('||')
                        combine_op = any
                    else:
                        condition_parts = [filter_part]
                        combine_op = all
                        
                    # Parse each condition
                    parsed_conditions = []
                    for part in condition_parts:
                        field, op, value = parse_filter_condition(part)
                        if field:
                            parsed_conditions.append((field, op, value))
                    
                    # Extract the field being aggregated and the operation type
                    reduce_part = formula.formula.split('.reduce')[1]
                    
                    # Determine the operation type
                    if 'count' in formula.formula.lower():
                        result = len(filtered_items)
                    else:
                        # Extract field name from reduce operation
                        sum_field = formula.formula.split('item.')[2].split(' ')[0].rstrip(')')
                        
                        if not filtered_items:
                            result = 0
                        elif 'min' in formula.formula.lower():
                            result = min(item.get(sum_field, 0) for item in filtered_items)
                        elif 'max' in formula.formula.lower():
                            result = max(item.get(sum_field, 0) for item in filtered_items)
                        elif 'avg' in formula.formula.lower():
                            result = sum(item.get(sum_field, 0) for item in filtered_items) / len(filtered_items)
                        else:  # sum
                            result = sum(item.get(sum_field, 0) for item in filtered_items)
                else:
                    # Get constants from _constants_data
                    constants = {}
                    if hasattr(self, '_constants_data') and self._constants_data:
                        try:
                            constants = json.loads(self._constants_data)
                            # Convert all constant values to float for calculations
                            constants = {k: flt(v) for k, v in constants.items()}
                        except Exception as e:
                            frappe.log_error(f"Error loading constants data: {str(e)}")
                    
                    # Create evaluation context for regular formulas
                    eval_globals = {
                        "items": items_data,
                        "frappe": frappe,
                        "math": math,
                        "flt": flt,
                        "cint": cint,
                        "sum": lambda field: sum(item.get(field, 0) for item in items_data),
                        "avg": lambda field: sum(item.get(field, 0) for item in items_data) / len(items_data) if items_data else 0,
                        "min": lambda field: min(item.get(field, 0) for item in items_data) if items_data else 0,
                        "max": lambda field: max(item.get(field, 0) for item in items_data) if items_data else 0,
                        "count": lambda field: sum(1 for item in items_data if field in item),
                        "distinct_count": lambda field: len(set(item.get(field) for item in items_data if field in item)),
                        "doc_totals": totals,  # Use the current totals
                        "constants": constants  # Add constants to context
                    }
                    
                    # Add constants directly to context for backward compatibility
                    eval_globals.update(constants)
                    
                    result = eval(formula.formula, eval_globals)
                
                # Convert result based on field type
                if formula.field_type == 'Int':
                    result = cint(result)
                else:  # Float, Currency, or Percent
                    result = flt(result)
                    
                totals[formula.field_name] = result
                
            except Exception as e:
                frappe.throw(f"Error calculating {formula.label}: {str(e)}")
        
        self._totals_data = json.dumps(totals)

    def get_item_variables(self, item):
        """Get all variables for an item with defaults"""
        if not item._data:
            # Get scope type for default values
            scope_type = frappe.get_doc("Scope Type", self.scope_type)
            defaults = {}
            
            # Initialize with default values from field configuration
            for field in scope_type.scope_fields:
                if field.default_value:
                    # Convert default value based on field type
                    if field.field_type in ['Float', 'Currency']:
                        defaults[field.field_name] = flt(field.default_value)
                    elif field.field_type == 'Int':
                        defaults[field.field_name] = cint(field.default_value)
                    elif field.field_type == 'Check':
                        defaults[field.field_name] = field.default_value.lower() == 'true'
                    else:
                        defaults[field.field_name] = field.default_value
            
            return defaults
            
        try:
            return json.loads(item._data)
        except json.JSONDecodeError:
            return {}

    def get_scope_totals(self):
        """Get scope-level totals with error handling"""
        if not self._totals_data:
            return {}
        try:
            return json.loads(self._totals_data)
        except json.JSONDecodeError:
            frappe.log_error("Invalid JSON in _totals_data")
            return {}

    def get_field_values(self, field_name):
        """Get all values for a field"""
        return [
            flt(item.get_dynamic_value(field_name), 0)
            for item in self.items
            if item.get_dynamic_value(field_name) is not None
        ]

@frappe.whitelist()
def get_scope_fields(scope_type):
    """Get field configurations for a scope type"""
    doc = frappe.get_doc("Scope Type", scope_type)
    return doc.scope_fields


@frappe.whitelist()
def save_scope_item(scope_items, item_data):
    """Save a scope item"""
    doc = frappe.get_doc("Scope Items", scope_items)

    # Parse item data
    item_data = frappe.parse_json(item_data)

    # Create or update item
    if item_data.get("row_id"):
        # Update existing item
        item = next((item for item in doc.items if item.row_id == item_data.get("row_id")), None)
        if item:
            item.item_name = item_data.get("item_name")
            # Handle other fields
            for field_name, value in item_data.items():
                if field_name not in ["row_id", "item_name"]:
                    item.set_dynamic_value(field_name, value)
    else:
        # Create new item
        item = doc.append("items", {
            "row_id": frappe.utils.random_string(10),
            "item_name": item_data.get("item_name")
        })
        # Handle other fields
        for field_name, value in item_data.items():
            if field_name not in ["row_id", "item_name"]:
                item.set_dynamic_value(field_name, value)

    doc.save()
    return doc.as_dict()


@frappe.whitelist()
def delete_scope_item(scope_items, row_id):
    """Delete a scope item"""
    doc = frappe.get_doc("Scope Items", scope_items)
    
    # Find and remove the item
    items_before = len(doc.items)
    doc.items = [item for item in doc.items if item.row_id != row_id]
    
    if len(doc.items) == items_before:
        frappe.throw(_("Item not found"))
        
    doc.save()
    return doc.as_dict()


@frappe.whitelist()
def save_multiple_scope_items(scope_items, items_data, clear_existing=False):
    """Save multiple scope items at once"""
    if isinstance(items_data, str):
        items_data = json.loads(items_data)
    
    # Convert clear_existing to boolean if it comes as string or integer
    clear_existing = bool(int(clear_existing)) if str(clear_existing).isdigit() else bool(clear_existing)
        
    doc = frappe.get_doc("Scope Items", scope_items)
    
    # If clear_existing is True, remove all existing items
    if clear_existing:
        # Create a new empty list for items
        doc.set('items', [])
    
    # Add new items
    for item_data in items_data:
        # Create new item as a dictionary first
        new_item = {
            'doctype': 'Scope Item',  # This is the child doctype name
            'parent': doc.name,
            'parenttype': 'Scope Items',
            'parentfield': 'items',
            'row_id': frappe.utils.cstr(frappe.utils.now_datetime().timestamp()),
            'item_name': item_data.get('item_name'),
            '_data': json.dumps(item_data)
        }
        
        # Append the item using the dictionary
        doc.append('items', new_item)
    
    # Save the document
    doc.save()
    frappe.db.commit()  # Ensure changes are committed
    
    return doc


@frappe.whitelist()
def get_template_with_formulas(scope_items, include_data=False):
    """Generate an Excel template with formulas for scope items"""
    # Convert include_data to boolean if it comes as string or integer
    include_data = bool(int(include_data)) if str(include_data).isdigit() else bool(include_data)
    
    doc = frappe.get_doc("Scope Items", scope_items)
    scope_type = frappe.get_doc("Scope Type", doc.scope_type)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Create Data sheet (rename the default sheet)
    data_sheet = wb.active
    data_sheet.title = 'Data'
    
    # Get field configurations
    scope_fields = scope_type.scope_fields
    manual_fields = [f.field_name for f in scope_fields if not f.auto_calculate]
    calculated_fields = [f for f in scope_fields if f.auto_calculate]
    headers = ['item_name'] + manual_fields + [f.field_name for f in calculated_fields]
    
    # Helper function to convert JS formula to Excel formula
    def convert_to_excel_formula(formula, row_num):
        if not formula:
            return ""
            
        # Create variable mappings
        var_map = {field: f"Data!{get_column_letter(i+1)}{row_num}" 
                  for i, field in enumerate(headers)}
        const_map = {key: f"Constants!B{i+2}" 
                    for i, key in enumerate(doc._constants_data and json.loads(doc._constants_data) or {})}
        totals_map = {key: f"Totals!B{i+2}" 
                     for i, key in enumerate(doc._totals_data and json.loads(doc._totals_data) or {})}
        
        # Replace variable references
        excel_formula = formula
        for var_name, excel_ref in var_map.items():
            excel_formula = excel_formula.replace(f"variables['{var_name}']", excel_ref)
        
        # Replace constants references
        for const_name, excel_ref in const_map.items():
            excel_formula = excel_formula.replace(f"constants['{const_name}']", excel_ref)
            excel_formula = excel_formula.replace(f"constants.{const_name}", excel_ref)
        
        # Replace doc_totals references
        for total_name, excel_ref in totals_map.items():
            excel_formula = excel_formula.replace(f"doc_totals['{total_name}']", excel_ref)
            excel_formula = excel_formula.replace(f"doc_totals.{total_name}", excel_ref)
        
        # Convert JavaScript operators to Excel
        excel_formula = (excel_formula
            .replace("?", "IF(")
            .replace(":", ",")
            .replace("&&", "*")
            .replace("||", "+")
            .replace("===", "=")
            .replace("==", "=")
            .replace(">=", ">=")
            .replace("<=", "<=")
            .replace("true", "1")
            .replace("false", "0"))
        
        # Fix nested IF statements
        while "))" in excel_formula:
            excel_formula = excel_formula.replace("))", ")")
        
        return excel_formula

    def convert_total_formula(formula):
        if not formula:
            return ""
            
        excel_formula = formula
        
        # Handle ternary operators first
        while '?' in excel_formula and ':' in excel_formula:
            # Find the innermost ternary operator
            start_idx = excel_formula.rindex('?')
            condition_start = excel_formula.rfind('(', 0, start_idx)
            if condition_start == -1:
                condition_start = 0
            condition = excel_formula[condition_start:start_idx].strip()
            
            # Find corresponding : and get true/false values
            colon_idx = excel_formula.find(':', start_idx)
            next_ternary = excel_formula.find('?', start_idx + 1)
            if next_ternary != -1 and next_ternary < colon_idx:
                # Nested ternary, skip this one
                continue
                
            end_idx = excel_formula.find(')', colon_idx)
            if end_idx == -1:
                end_idx = len(excel_formula)
            
            true_value = excel_formula[start_idx + 1:colon_idx].strip()
            false_value = excel_formula[colon_idx + 1:end_idx].strip()
            
            # Convert to Excel IF
            replacement = f"IF({condition}, {true_value}, {false_value})"
            excel_formula = excel_formula[:condition_start] + replacement + excel_formula[end_idx:]
        
        # Handle multiple condition pricing (switch-case like structures)
        if 'let' in excel_formula and '===' in excel_formula:
            lines = excel_formula.split('\n')
            conditions = []
            for line in lines:
                if '===' in line and '?' in line and ':' in line:
                    condition = line.split('===')[0].split('variables[')[1].split(']')[0].strip("'")
                    value = line.split('===')[1].split('?')[0].strip()
                    result = line.split('?')[1].split(':')[0].strip()
                    
                    col_letter = get_column_letter(headers.index(condition) + 1)
                    conditions.append(f"IF(Data!{col_letter}2:{col_letter}1000={value}, {result}")
            
            if conditions:
                # Add the final else value
                else_value = lines[-1].strip()
                if else_value.isdigit() or else_value.replace('.', '').isdigit():
                    excel_formula = ', '.join(conditions) + f", {else_value}" + (")" * len(conditions))
        
        # Handle Math functions
        excel_formula = excel_formula.replace('Math.PI', '3.14159')
        excel_formula = excel_formula.replace('Math.pow(', 'POWER(')
        excel_formula = excel_formula.replace('Math.sqrt(', 'SQRT(')
        excel_formula = excel_formula.replace('Math.abs(', 'ABS(')
        excel_formula = excel_formula.replace('Math.round(', 'ROUND(')
        excel_formula = excel_formula.replace('Math.floor(', 'FLOOR(')
        excel_formula = excel_formula.replace('Math.ceil(', 'CEILING(')
        
        # Handle filtered aggregates
        if 'items.filter' in excel_formula:
            # Extract filter conditions
            filter_part = excel_formula.split('items.filter(')[1].split('.reduce')[0]
            
            # Handle multiple conditions
            if '&&' in filter_part:
                condition_parts = filter_part.split('&&')
                combine_op = '*'  # AND becomes multiplication in Excel
            elif '||' in filter_part:
                condition_parts = filter_part.split('||')
                combine_op = '+'  # OR becomes addition in Excel
            else:
                condition_parts = [filter_part]
                combine_op = '*'
            
            # Convert each condition to Excel
            conditions = []
            for part in condition_parts:
                if '=>' in part:
                    condition = part.split('=>')[1].strip()
                    # Convert condition to Excel format
                    condition = condition.replace('item.', '')
                    condition = condition.replace('===', '=')
                    condition = condition.replace('==', '=')
                    conditions.append(f"IF({condition}, 1, 0)")
            
            # Extract the field being summed
            reduce_part = excel_formula.split('.reduce')[1]
            if 'item.' in reduce_part:
                field = reduce_part.split('item.')[1].split(' ')[0].rstrip(')')
                col_letter = get_column_letter(headers.index(field) + 1)
                
                # Create SUMIFS formula
                conditions_str = combine_op.join(conditions)
                excel_formula = f"SUMPRODUCT({col_letter}2:{col_letter}1000, {conditions_str})"
        else:
            # Handle standard aggregate functions
            for field in headers:
                col_letter = get_column_letter(headers.index(field) + 1)
                
                # Standard aggregates
                excel_formula = excel_formula.replace(
                    f"sum('{field}')",
                    f"SUM(Data!{col_letter}2:{col_letter}1000)"
                )
                excel_formula = excel_formula.replace(
                    f"count('{field}')",
                    f"COUNTIF(Data!{col_letter}2:{col_letter}1000,\"<>\")"
                )
                excel_formula = excel_formula.replace(
                    f"avg('{field}')",
                    f"AVERAGE(Data!{col_letter}2:{col_letter}1000)"
                )
                excel_formula = excel_formula.replace(
                    f"min('{field}')",
                    f"MIN(Data!{col_letter}2:{col_letter}1000)"
                )
                excel_formula = excel_formula.replace(
                    f"max('{field}')",
                    f"MAX(Data!{col_letter}2:{col_letter}1000)"
                )
                excel_formula = excel_formula.replace(
                    f"distinct_count('{field}')",
                    f"SUMPRODUCT((Data!{col_letter}2:{col_letter}1000<>\"\")/COUNTIF(Data!{col_letter}2:{col_letter}1000,Data!{col_letter}2:{col_letter}1000&\"*\"))"
                )
                
                # Handle variables references
                excel_formula = excel_formula.replace(
                    f"variables['{field}']",
                    f"Data!{col_letter}2:{col_letter}1000"
                )
            
            # Handle doc_totals references
            for i, formula in enumerate(totals_config, 2):
                excel_formula = excel_formula.replace(
                    f"doc_totals['{formula.field_name}']",
                    f"Totals!B{i}"
                )
                excel_formula = excel_formula.replace(
                    f"doc_totals.{formula.field_name}",
                    f"Totals!B{i}"
                )
            
            # Handle constants
            constants = json.loads(doc._constants_data) if doc._constants_data else {}
            for i, (key, value) in enumerate(constants.items(), 2):
                excel_formula = excel_formula.replace(
                    f"constants['{key}']",
                    f"Constants!B{i}"
                )
                excel_formula = excel_formula.replace(
                    f"constants.{key}",
                    f"Constants!B{i}"
                )
        
        # Replace Python/JavaScript operators with Excel operators
        excel_formula = excel_formula.replace("&&", "*")  # Logical AND
        excel_formula = excel_formula.replace("||", "+")  # Logical OR
        excel_formula = excel_formula.replace("===", "=")
        excel_formula = excel_formula.replace("==", "=")
        excel_formula = excel_formula.replace(">=", ">=")
        excel_formula = excel_formula.replace("<=", "<=")
        excel_formula = excel_formula.replace("!==", "<>")
        excel_formula = excel_formula.replace("!=", "<>")
        
        return excel_formula

    # Set up Data sheet
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
        
        # Auto-adjust column width based on header content
        column_letter = get_column_letter(col)
        data_sheet.column_dimensions[column_letter].width = max(len(str(header)) + 4, 12)
    
    # Freeze the header row
    data_sheet.freeze_panes = "A2"
    
    start_row = 2
    # Add existing data if requested
    if include_data and doc.items:
        for item in doc.items:
            try:
                data = json.loads(item._data or '{}')
            except json.JSONDecodeError:
                data = {}
            
            # Add item data
            for col, field_name in enumerate(headers, 1):
                cell = data_sheet.cell(row=start_row, column=col)
                if field_name == 'item_name':
                    cell.value = item.item_name
                else:
                    cell.value = data.get(field_name)
                    
                # Add formula for calculated fields
                field = next((f for f in scope_fields if f.field_name == field_name and f.auto_calculate), None)
                if field:
                    cell.value = f"={convert_to_excel_formula(field.calculation_formula, start_row)}"
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")
            start_row += 1
    
    # Add formula rows
    for row in range(start_row, start_row + 50):  # Add 50 rows after existing data
        for col, field_name in enumerate(headers, 1):
            cell = data_sheet.cell(row=row, column=col)
            field = next((f for f in scope_fields if f.field_name == field_name), None)
            
            if field and field.auto_calculate:
                cell.value = f"={convert_to_excel_formula(field.calculation_formula, row)}"
                cell.fill = PatternFill("solid", fgColor="F5F5F5")  # Light gray background for computed cells

    # Create Constants sheet
    constants_sheet = wb.create_sheet('Constants')
    constants_sheet.cell(row=1, column=1, value='Name').font = Font(bold=True)
    constants_sheet.cell(row=1, column=2, value='Value').font = Font(bold=True)
    
    if doc._constants_data:
        constants = json.loads(doc._constants_data)
        for i, (key, value) in enumerate(constants.items(), 2):
            constants_sheet.cell(row=i, column=1, value=key)
            constants_sheet.cell(row=i, column=2, value=value)
    
    # Create Totals sheet
    totals_sheet = wb.create_sheet('Totals')
    totals_sheet.cell(row=1, column=1, value='Name').font = Font(bold=True)
    totals_sheet.cell(row=1, column=2, value='Formula').font = Font(bold=True)
    
    # Get totals data and formulas
    scope_type_doc = frappe.get_doc("Scope Type", doc.scope_type)
    totals_config = scope_type_doc.calculation_formulas if hasattr(scope_type_doc, 'calculation_formulas') else []
    
    # Get all items data for formula calculations
    items_data = []
    for item in doc.items:
        try:
            item_data = json.loads(item._data) if item._data else {}
            items_data.append(item_data)
        except json.JSONDecodeError:
            items_data.append({})
    
    # Calculate totals using the same logic as in calculate_totals
    totals = {}
    eval_globals = {
        "items": items_data,
        "frappe": frappe,
        "math": math,
        "flt": flt,
        "cint": cint,
        "sum": lambda field: sum(item.get(field, 0) for item in items_data),
        "avg": lambda field: sum(item.get(field, 0) for item in items_data) / len(items_data) if items_data else 0,
        "min": lambda field: min(item.get(field, 0) for item in items_data) if items_data else 0,
        "max": lambda field: max(item.get(field, 0) for item in items_data) if items_data else 0,
        "count": lambda field: sum(1 for item in items_data if field in item),
        "distinct_count": lambda field: len(set(item.get(field) for item in items_data if field in item)),
        "doc_totals": totals,
        "constants": json.loads(doc._constants_data) if doc._constants_data else {}
    }
    
    for i, formula in enumerate(totals_config, 2):
        totals_sheet.cell(row=i, column=1, value=formula.label)
        excel_formula = convert_total_formula(formula.formula)
        totals_sheet.cell(row=i, column=2, value=f"={excel_formula}")
    
    # Save workbook to a temporary file
    temp_path = os.path.join(frappe.get_site_path(), 'private', 'files', 'scope_items_template.xlsx')
    wb.save(temp_path)
    
    # Save the file to Frappe's file system
    with open(temp_path, 'rb') as f:
        file_content = f.read()
        
    file_doc = save_file(
        fname='scope_items_template.xlsx',
        content=file_content,
        dt='Scope Items',
        dn=scope_items,
        folder='Home',
        is_private=1
    )
    
    # Clean up temporary file
    os.remove(temp_path)
    
    return {
        'file_url': file_doc.file_url,
        'file_name': 'scope_items_template.xlsx'
    }
