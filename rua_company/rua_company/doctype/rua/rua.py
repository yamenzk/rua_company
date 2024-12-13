# Copyright (c) 2024, Yamen Zakhour and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from frappe.utils import get_site_base_path, get_files_path
import os
from datetime import datetime
import base64
from PIL import Image as PILImage
import io

class Rua(Document):
    pass

def get_company_info():
    """Get company information from Rua settings"""
    return frappe.db.get_singles_dict('Rua')

def define_styles():
    """Define consistent styles for the workbook"""
    styles = {
        'header': {
            'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
        },
        'subheader': {
            'font': Font(name='Calibri', size=11, bold=True, color='2C3E50'),
            'fill': PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='B3B3B3'),
                right=Side(style='thin', color='B3B3B3'),
                top=Side(style='thin', color='B3B3B3'),
                bottom=Side(style='thin', color='B3B3B3')
            )
        },
        'data': {
            'font': Font(name='Calibri', size=10),
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # Added fill
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin', color='E5E5E5'),
                right=Side(style='thin', color='E5E5E5'),
                top=Side(style='thin', color='E5E5E5'),
                bottom=Side(style='thin', color='E5E5E5')
            )
        },
        'total': {
            'font': Font(name='Calibri', size=11, bold=True, color='2C3E50'),
            'fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'border': Border(
                left=Side(style='thin', color='B3B3B3'),
                right=Side(style='thin', color='B3B3B3'),
                top=Side(style='thin', color='B3B3B3'),
                bottom=Side(style='double', color='B3B3B3')
            )
        }
    }
    return styles

def apply_modern_branding(ws, company_info, report_title, from_date, to_date):
    """Apply clean, modern branding with full-width header"""
    # Get brand color with fallback
    brand_color = company_info.get('brand_color', '#2C3E50').lstrip('#')
    
    # Set column widths
    column_widths = {
        'A': 8,   # S.No
        'B': 15,  # Date
        'C': 15,  # Voucher
        'D': 25,  # Names
        'E': 25,  # TRN
        'F': 15,  # Amount
        'G': 15,  # Tax
        'H': 15,  # Total
        'I': 12   # Emirate
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    current_row = 1
    
    # Logo Row
    if company_info.get('logo_horizontal'):
        logo_path = frappe.get_site_path('public', company_info.logo_horizontal.lstrip('/'))
        if os.path.exists(logo_path):
            with PILImage.open(logo_path) as img:
                img = img.convert('RGB')
                max_height = 45  # pixels
                aspect_ratio = img.width / img.height
                new_width = int(max_height * aspect_ratio)
                img = img.resize((new_width, max_height))
                
                img_bytes = io.BytesIO()
                img.save(img_bytes, format='PNG')
                img_bytes.seek(0)
                
                logo = Image(img_bytes)
                logo.width = new_width
                logo.height = max_height
                ws.add_image(logo, 'A1')
        
        # Merge cells C to I and set background color
        ws.merge_cells(f'C{current_row}:I{current_row}')
        for col in range(3, 10):  # C through I (3 to 9 in openpyxl)
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
                
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # Company Name - Large and prominent
    ws.row_dimensions[current_row].height = 35
    company_cell = ws.cell(row=current_row, column=1)  # Start from column A
    company_cell.value = company_info.get('company_name', '')
    company_cell.font = Font(name='Calibri', size=18, bold=True, color=brand_color)
    company_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    # TRN/VAT info - Right aligned in the same row
    trn_cell = ws.cell(row=current_row, column=9)  # Adjusted column for better spacing
    trn_cell.value = f"TRN: {company_info.get('trn', '')}  |  VAT: {company_info.get('vat', '5')}%"
    trn_cell.font = Font(name='Calibri', size=11, color='666666')
    trn_cell.alignment = Alignment(horizontal='right', vertical='center')
    current_row += 1

    # Separator line with brand color
    ws.row_dimensions[current_row].height = 4
    for col in range(1, 10):  # A through I
        cell = ws.cell(row=current_row, column=col)
        cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
    current_row += 1

    # Report Title and Date Range
    ws.row_dimensions[current_row].height = 30
    
    # Report Title - Left aligned
    title_cell = ws.cell(row=current_row, column=1)  # Start from column A
    title_cell.value = report_title
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='2C3E50')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    
    # Date Range - Right aligned
    date_cell = ws.cell(row=current_row, column=9)  # Adjusted column for better spacing
    date_cell.value = f"Period: {from_date} to {to_date}"
    date_cell.font = Font(name='Calibri', size=11, color='666666')
    date_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    current_row += 2  # Add space before table headers
    
    return current_row

def style_modern_table(ws, start_row, headers, styles):
    """Style table headers with a clean, modern look"""
    ws.row_dimensions[start_row].height = 30
    
    header_font = Font(name='Calibri', size=11, bold=True, color='2C3E50')
    header_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    header_border = Border(
        bottom=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    return start_row + 1

def apply_row_styling(ws, row, data, styles, is_striped=True):
    """Apply modern row styling with subtle stripes"""
    stripe_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    
    for col, (value, align) in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        
        # Apply base styles
        cell.font = Font(name='Calibri', size=10, color='2C3E50')
        cell.border = Border(
            bottom=Side(style='thin', color='E5E5E5')
        )
        
        # Apply stripe pattern
        if is_striped and row % 2 == 0:
            cell.fill = stripe_fill
            
        # Set alignment
        cell.alignment = Alignment(horizontal=align, vertical='center')
        
        # Format based on column type
        if col == 1:  # S.No
            cell.number_format = '0'
        elif col == 2:  # Date
            cell.number_format = 'DD/MM/YYYY'
        elif col in [6, 7, 8]:  # Payment Amount, Tax Amount, Total Amount
            cell.number_format = '"AED "#,##0.00'
        
        # Format numbers
        if isinstance(value, (int, float)):
            pass

def apply_style(cell, style):
    """Apply a predefined style to a cell"""
    # Apply each style property if it exists
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']

def adjust_color_brightness(hex_color, factor):
    """Adjust the brightness of a hex color"""
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    new_rgb = tuple(min(255, int(c * factor)) for c in rgb)
    return '%02x%02x%02x' % new_rgb

def get_filtered_vouchers(doctype, filters, include_no_trn):
    """Get vouchers filtered by TRN if needed"""
    vouchers = frappe.get_all(
        doctype,
        filters=filters,
        fields=["name", "date", "party", "amount", "trn", "emirate"]
    )
    
    if not int(include_no_trn):
        vouchers = [v for v in vouchers if v.trn not in [None, "", "0"]]
    
    return vouchers

@frappe.whitelist()
def generate_vat_report(from_date, to_date, include_no_trn=1):
    """Generate VAT report with enhanced styling"""
    company_info = get_company_info()
    vat_percentage = float(company_info.get('vat', '5'))
    styles = define_styles()
    
    wb = openpyxl.Workbook()
    generate_supplier_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn)
    generate_customer_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn)
    generate_petty_cash_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn)
    generate_summary_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn)
    
    # Save to temporary file
    file_name = f'VAT_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    temp_file_path = os.path.join(get_files_path(), file_name)
    wb.save(temp_file_path)
    
    # Read and encode file
    with open(temp_file_path, 'rb') as f:
        file_content = f.read()
    
    # Clean up and return
    os.remove(temp_file_path)
    return {
        'file_name': file_name,
        'file_content': base64.b64encode(file_content).decode('utf-8')
    }

def generate_supplier_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn):
    """Generate supplier sheet with enhanced styling"""
    ws = wb.active
    ws.title = "Supplier"
    
    headers = ['S. No.', 'Payment Date', 'Voucher Number', 'Supplier Name', 'Supplier TRN',
               'Payment Amount', 'Tax Amount', 'Total Amount', 'Emirate']
    
    # Apply branding and get starting row
    start_row = apply_modern_branding(ws, company_info, "Supplier VAT Report", from_date, to_date)
    
    # Style table and get data starting row
    data_start_row = style_modern_table(ws, start_row, headers, styles)
    
    # Get payment vouchers
    payment_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "type": "Pay",
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 0
        },
        include_no_trn
    )
    
    # Fill data
    row = data_start_row
    for idx, pv in enumerate(payment_vouchers, 1):
        amount = float(pv.amount)
        base_amount = amount / (1 + (vat_percentage / 100))
        tax_amount = amount - base_amount
        
        data = [
            (idx, 'center'),
            (pv.date, 'center'),
            (pv.name, 'left'),
            (pv.party, 'left'),
            (pv.trn, 'center'),
            (round(base_amount, 2), 'right'),
            (round(tax_amount, 2), 'right'),
            (amount, 'right'),
            (pv.emirate, 'center')
        ]
        
        apply_row_styling(ws, row, data, styles)
        row += 1
    
    # Add totals
    if payment_vouchers:
        row += 1
        total_base = sum(float(pv.amount) / (1 + (vat_percentage / 100)) for pv in payment_vouchers)
        total_tax = sum(float(pv.amount) - (float(pv.amount) / (1 + (vat_percentage / 100))) for pv in payment_vouchers)
        total_amount = sum(float(pv.amount) for pv in payment_vouchers)
        
        # Merge cells for total label
        ws.merge_cells(f'A{row}:E{row}')
        total_label = ws.cell(row=row, column=1, value="Totals")
        apply_style(total_label, styles['total'])
        
        # Add total amounts
        for value, col in [
            (round(total_base, 2), 6),
            (round(total_tax, 2), 7),
            (total_amount, 8)
        ]:
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, styles['total'])
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')

def generate_customer_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn):
    """Generate customer sheet with enhanced styling"""
    ws = wb.create_sheet(title="Customer")
    
    headers = ['S. No.', 'Payment Date', 'Voucher Number', 'Customer Name', 'Customer TRN',
               'Payment Amount', 'Tax Amount', 'Total Amount', 'Emirate']
    
    # Apply branding and get starting row
    start_row = apply_modern_branding(ws, company_info, "Customer VAT Report", from_date, to_date)
    
    # Style table and get data starting row
    data_start_row = style_modern_table(ws, start_row, headers, styles)
    
    # Get payment vouchers
    payment_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "type": "Receive",
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 0
        },
        include_no_trn
    )
    
    # Fill data
    row = data_start_row
    for idx, pv in enumerate(payment_vouchers, 1):
        amount = float(pv.amount)
        base_amount = amount / (1 + (vat_percentage / 100))
        tax_amount = amount - base_amount
        
        data = [
            (idx, 'center'),
            (pv.date, 'center'),
            (pv.name, 'left'),
            (pv.party, 'left'),
            (pv.trn, 'center'),
            (round(base_amount, 2), 'right'),
            (round(tax_amount, 2), 'right'),
            (amount, 'right'),
            (pv.emirate, 'center')
        ]
        
        apply_row_styling(ws, row, data, styles)
        row += 1
    
    # Add totals
    if payment_vouchers:
        row += 1
        total_base = sum(float(pv.amount) / (1 + (vat_percentage / 100)) for pv in payment_vouchers)
        total_tax = sum(float(pv.amount) - (float(pv.amount) / (1 + (vat_percentage / 100))) for pv in payment_vouchers)
        total_amount = sum(float(pv.amount) for pv in payment_vouchers)
        
        # Merge cells for total label
        ws.merge_cells(f'A{row}:E{row}')
        total_label = ws.cell(row=row, column=1, value="Totals")
        apply_style(total_label, styles['total'])
        
        # Add total amounts
        for value, col in [
            (round(total_base, 2), 6),
            (round(total_tax, 2), 7),
            (total_amount, 8)
        ]:
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, styles['total'])
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')

def generate_petty_cash_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn):
    """Generate petty cash sheet with enhanced styling"""
    ws = wb.create_sheet(title="Petty Cash")
    
    headers = ['S. No.', 'Payment Date', 'Voucher Number', 'Party Name', 'Party TRN',
               'Payment Amount', 'Tax Amount', 'Total Amount', 'Emirate']
    
    # Apply branding and get starting row
    start_row = apply_modern_branding(ws, company_info, "Petty Cash VAT Report", from_date, to_date)
    
    # Style table and get data starting row
    data_start_row = style_modern_table(ws, start_row, headers, styles)
    
    # Get payment vouchers
    payment_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 1
        },
        include_no_trn
    )
    
    # Fill data
    row = data_start_row
    for idx, pv in enumerate(payment_vouchers, 1):
        amount = float(pv.amount)
        base_amount = amount / (1 + (vat_percentage / 100))
        tax_amount = amount - base_amount
        
        data = [
            (idx, 'center'),
            (pv.date, 'center'),
            (pv.name, 'left'),
            (pv.party, 'left'),
            (pv.trn, 'center'),
            (round(base_amount, 2), 'right'),
            (round(tax_amount, 2), 'right'),
            (amount, 'right'),
            (pv.emirate, 'center')
        ]
        
        apply_row_styling(ws, row, data, styles)
        row += 1
    
    # Add totals
    if payment_vouchers:
        row += 1
        total_base = sum(float(pv.amount) / (1 + (vat_percentage / 100)) for pv in payment_vouchers)
        total_tax = sum(float(pv.amount) - (float(pv.amount) / (1 + (vat_percentage / 100))) for pv in payment_vouchers)
        total_amount = sum(float(pv.amount) for pv in payment_vouchers)
        
        # Merge cells for total label
        ws.merge_cells(f'A{row}:E{row}')
        total_label = ws.cell(row=row, column=1, value="Totals")
        apply_style(total_label, styles['total'])
        
        # Add total amounts
        for value, col in [
            (round(total_base, 2), 6),
            (round(total_tax, 2), 7),
            (total_amount, 8)
        ]:
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, styles['total'])
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')

def generate_summary_sheet(wb, from_date, to_date, vat_percentage, company_info, styles, include_no_trn):
    """Generate summary sheet with party-wise totals and final summary"""
    ws = wb.create_sheet(title="Summary")
    
    # Apply branding
    current_row = apply_modern_branding(ws, company_info, "VAT Summary Report", from_date, to_date)
    current_row += 1
    
    # Helper function to get party-wise totals
    def get_party_totals(vouchers):
        totals = {}
        for pv in vouchers:
            amount = float(pv.amount)
            base_amount = amount / (1 + (vat_percentage / 100))
            tax_amount = amount - base_amount
            
            if pv.party not in totals:
                totals[pv.party] = {'base': 0, 'tax': 0, 'total': 0, 'trn': pv.trn}
            
            totals[pv.party]['base'] += base_amount
            totals[pv.party]['tax'] += tax_amount
            totals[pv.party]['total'] += amount
        
        return totals
    
    # Function to create a summary table
    def create_summary_table(title, data, start_row):
        # Add title
        title_cell = ws.cell(row=start_row, column=3, value=title)  # Start from column C
        title_cell.font = Font(name='Calibri', size=12, bold=True, color='2C3E50')
        start_row += 1
        
        # Add headers
        headers = ['Party Name', 'TRN', 'Base Amount', 'Tax Amount', 'Total Amount']
        for col, header in enumerate(headers, 3):  # Start from column C (3)
            cell = ws.cell(row=start_row, column=col, value=header)
            apply_style(cell, styles['header'])
        start_row += 1
        
        # Add data
        total_base = total_tax = total_amount = 0
        for party, values in data.items():
            ws.cell(row=start_row, column=3, value=party).alignment = Alignment(horizontal='left')  # Column C
            ws.cell(row=start_row, column=4, value=values['trn']).alignment = Alignment(horizontal='center')  # Column D
            
            base_cell = ws.cell(row=start_row, column=5, value=round(values['base'], 2))  # Column E
            tax_cell = ws.cell(row=start_row, column=6, value=round(values['tax'], 2))   # Column F
            total_cell = ws.cell(row=start_row, column=7, value=round(values['total'], 2))  # Column G
            
            for cell in [base_cell, tax_cell, total_cell]:
                cell.number_format = '"AED "#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            
            total_base += values['base']
            total_tax += values['tax']
            total_amount += values['total']
            start_row += 1
        
        # Add totals
        start_row += 1
        ws.merge_cells(f'C{start_row}:D{start_row}')  # Merge C-D instead of A-B
        total_label = ws.cell(row=start_row, column=3, value="Total")  # Column C
        apply_style(total_label, styles['total'])
        
        for col, value in [(5, total_base), (6, total_tax), (7, total_amount)]:  # Adjusted columns
            cell = ws.cell(row=start_row, column=col, value=round(value, 2))
            apply_style(cell, styles['total'])
            cell.number_format = '"AED "#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        return start_row + 2, total_tax
    
    # Get all vouchers
    supplier_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "type": "Pay",
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 0
        },
        include_no_trn
    )
    
    customer_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "type": "Receive",
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 0
        },
        include_no_trn
    )
    
    petty_cash_vouchers = get_filtered_vouchers(
        "Payment Voucher",
        {
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "is_petty_cash": 1
        },
        include_no_trn
    )
    
    # Create summary tables
    current_row, supplier_tax = create_summary_table("Supplier Summary", get_party_totals(supplier_vouchers), current_row)
    current_row, customer_tax = create_summary_table("Customer Summary", get_party_totals(customer_vouchers), current_row)
    current_row, petty_cash_tax = create_summary_table("Petty Cash Summary", get_party_totals(petty_cash_vouchers), current_row)
    
    # Final Summary
    current_row += 1
    ws.merge_cells(f'C{current_row}:G{current_row}')  # Merge C-G instead of A-E
    final_title = ws.cell(row=current_row, column=3, value="Final Summary")  # Column C
    final_title.font = Font(name='Calibri', size=14, bold=True, color='2C3E50')
    final_title.alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Add final calculations
    total_purchases_tax = supplier_tax + petty_cash_tax
    net_payable = total_purchases_tax - customer_tax
    
    summary_data = [
        ("Purchases (Including Petty Cash)", total_purchases_tax),
        ("Sales", customer_tax),
        ("Net Payable", net_payable)
    ]
    
    for label, value in summary_data:
        ws.merge_cells(f'C{current_row}:F{current_row}')  # Merge C-F instead of A-D
        label_cell = ws.cell(row=current_row, column=3, value=label)  # Column C
        value_cell = ws.cell(row=current_row, column=7, value=value)  # Column G
        
        label_cell.font = Font(name='Calibri', size=11, bold=True)
        value_cell.font = Font(name='Calibri', size=11, bold=True)
        value_cell.number_format = '"AED "#,##0.00'
        value_cell.alignment = Alignment(horizontal='right')
        
        current_row += 1
    
    # Adjust column widths
    column_widths = {'C': 30, 'D': 15, 'E': 15, 'F': 15, 'G': 15}  # Adjusted columns
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width